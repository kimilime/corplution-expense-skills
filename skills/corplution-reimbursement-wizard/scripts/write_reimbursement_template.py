#!/usr/bin/env python3
"""Write confirmed expense allocations into the reimbursement Excel workbook."""

from __future__ import annotations

import argparse
import hashlib
import copy
import json
import re
import sys
from collections import OrderedDict, defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover - Python < 3.11 fallback
    import tomli as tomllib


C = {
    "summary": "\u6c47\u603b",
    "local": "\u672c\u5730",
    "trip": "\u51fa\u5dee",
    "substitute": "\uff08\u62b5\uff09",
    "trip_meal": "\u51fa\u5dee\u9910\u8d39",
    "overtime_meal": "\u52a0\u73ed\u9910\u8d39",
    "business_trip_meal_policy": "\u51fa\u5dee\u9910\u8d39",
    "local_overtime_meal_policy": "\u672c\u5730\u52a0\u73ed\u9910\u8d39",
    "hotel_policy_first_tier": "\u9152\u5e97\uff08\u5317\u4e0a\u5e7f\u6df1\uff09",
    "hotel_policy_other_city": "\u9152\u5e97\uff08\u5176\u4ed6\u57ce\u5e02\uff09",
    "invoice_amount": "\u53d1\u7968\u91d1\u989d",
    "reimbursable_amount": "\u5b9e\u9645\u62a5\u9500",
    "meal_cap_ok": "\u672a\u8d85\u6807",
    "meal_cap_over_with_attendees": "\u8d85\u6807\uff0c\u5df2\u6709\u591a\u4eba\u4fe1\u606f\uff0c\u4ec5\u63d0\u793a\u590d\u6838",
    "meal_cap_over_needs_confirmation": "\u8d85\u6807\uff0c\u9700\u786e\u8ba4\u65e5\u671f/\u591a\u4eba/\u5b9e\u62a5\u91d1\u989d",
    "hotel_cap_missing_nights": "\u7f3a\u5c11\u665a\u6570\uff0c\u9700\u786e\u8ba4\u5165\u4f4f/\u79bb\u5e97/\u665a\u6570",
    "hotel_cap_missing_city": "\u7f3a\u5c11\u57ce\u5e02\uff0c\u9700\u786e\u8ba4\u57ce\u5e02\u6863\u4f4d",
    "hotel_cap_over_with_shared_room": "\u8d85\u6807\uff0c\u5df2\u6709\u540c\u4f4f/\u6807\u95f4\u4fe1\u606f\uff0c\u4ec5\u63d0\u793a\u590d\u6838",
    "hotel_cap_over_needs_confirmation": "\u8d85\u6807\uff0c\u9700\u786e\u8ba4\u665a\u6570/\u540c\u4f4f/\u5b9e\u62a5\u91d1\u989d",
}

# Policy numbers and year-coded values come from assets/policy.toml via
# policy_config; edit that file (not this one) when company policy changes.
from policy_config import load_policy
import evidence_paths
from exit_codes import ExitCode
import integrity
from io_utils import configure_utf8_stdio as configure_stdio
from json_io import read_json_object as load_json
import text_safety
from text_utils import normalize_text as clean
import subagent_protocol
import time_utils
from travel_ticket_utils import (
    contains_raw_ticket_evidence,
    is_flight_ticket,
    is_rail_ticket,
    route_from_text,
    ticket_note,
)
import value_utils

_POLICY = load_policy()
BUSINESS_TRIP_MEAL_DAILY_CAP = _POLICY.business_trip_meal_daily_cap
LOCAL_OVERTIME_MEAL_DAILY_CAP = _POLICY.local_overtime_meal_daily_cap
FIRST_TIER_HOTEL_CAP = _POLICY.first_tier_hotel_cap
OTHER_CITY_HOTEL_CAP = _POLICY.other_city_hotel_cap
ADMIN_CODE = _POLICY.admin_code
ADMIN_FALLBACK_CLIENT = _POLICY.admin_fallback_client
MOBILE_CLIENT = _POLICY.mobile_client
FIRST_TIER_CITIES = _POLICY.first_tier_cities

TRIP_MEAL_CONTEXTS = {"travel", "business_trip", "station_airport"}
OVERTIME_MEAL_CONTEXTS = {"overtime"}
MEAL_POLICY_NON_TRIGGERS = ["city", "amount_column", "final_template_column", "expenses_nature"]


AMOUNT_COLUMNS = {
    "hotel": "G",
    "travel": "H",
    "taxi": "I",
    "meal": "J",
    "mobile": "K",
    "other": "L",
}

ALLOWED_SOURCE_CATEGORIES = {"hotel", "travel", "taxi", "meal", "mobile", "other"}

ALLOCATION_TEXT_FIELDS = {
    "attendees",
    "approval_file",
    "business_reason",
    "city",
    "client_charge_code",
    "client_name",
    "correction_note",
    "destination",
    "destination_place_type",
    "expense_date",
    "expense_note",
    "expenses_nature",
    "final_note",
    "hotel_city",
    "journey_chain_match_reason",
    "journey_chain_route",
    "origin",
    "origin_place_type",
    "room_share_note",
    "room_shared_with",
    "route",
    "substitute_for",
}
WORKBOOK_ROW_TEXT_FIELDS = {
    "date",
    "requester",
    "client",
    "client_charge_code",
    "expenses_nature",
    "note",
}


PROOF_ORDER = {
    "flight": 1,
    "rail": 1,
    "railway": 1,
    "railway_e_ticket": 1,
    "hotel": 2,
    "taxi_didi": 3,
    "taxi": 3,
    "didi": 3,
    "gaode": 4,
    "meal": 5,
    "mobile": 6,
    "other": 7,
    "travel": 8,
    "unknown": 99,
}


ROW_ORDER = {
    "flight": 1,
    "rail": 1,
    "railway": 1,
    "railway_e_ticket": 1,
    "hotel": 2,
    "taxi_didi": 3,
    "taxi": 3,
    "didi": 3,
    "gaode": 4,
    "meal": 5,
    "mobile": 6,
    "other": 7,
    "travel": 8,
}


def bundled_template_path() -> Path:
    return Path(__file__).resolve().parents[1] / "assets" / "reimbursement-template.xlsx"


def bundled_layout_path() -> Path:
    return Path(__file__).resolve().parents[1] / "assets" / "reimbursement-workbook-layout.toml"


def resolve_template_arg(value: str | None) -> Path | None:
    if not value:
        return None
    if value.lower() == "bundled":
        return bundled_template_path()
    return Path(value)


def resolve_layout_arg(value: str | None) -> Path:
    return Path(value) if value else bundled_layout_path()


def load_layout(path: Path) -> dict[str, Any]:
    with path.open("rb") as handle:
        layout = tomllib.load(handle)
    required_sections = [
        "sheet",
        "columns",
        "rows",
        "fonts",
        "formats",
        "colors",
        "styles",
        "sample_rows",
        "labels",
        "instruction_row",
        "header_row",
    ]
    missing = [section for section in required_sections if section not in layout]
    if missing:
        raise ValueError(f"Workbook layout is missing sections: {', '.join(missing)}")
    expected_columns = int(layout["columns"]["count"])
    for row_name in ("instruction_row", "header_row"):
        values = layout[row_name].get("values", [])
        if len(values) != expected_columns:
            raise ValueError(f"{row_name}.values must contain {expected_columns} entries, got {len(values)}")
    return layout


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def workbook_text_issues(path: Path) -> list[str]:
    """Verify that saved cells did not suffer a terminal/encoding conversion."""
    findings: list[str] = []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str) or cell.value.startswith("="):
                        continue
                    findings.extend(
                        text_safety.find_suspect_text(
                            cell.value,
                            path=f"workbook.{worksheet.title}!{cell.coordinate}",
                            limit=max(1, 20 - len(findings)),
                        )
                    )
                    if len(findings) >= 20:
                        return findings
    finally:
        workbook.close()
    return findings


def normalized_client_name(unit: dict[str, Any]) -> str:
    client = clean(unit.get("client_name"))
    code = clean(unit.get("client_charge_code")).upper()
    if code != ADMIN_CODE:
        return client
    placeholder = client.lower() in {"", "admin", ADMIN_CODE.lower()}
    if unit.get("source_category") == "mobile" or unit.get("final_template_column") == "mobile":
        return MOBILE_CLIENT if placeholder or client == ADMIN_FALLBACK_CLIENT else client
    return ADMIN_FALLBACK_CLIENT if placeholder else client


def mobile_accounting_errors(unit: dict[str, Any]) -> list[str]:
    source_category = clean(unit.get("source_category"))
    final_column = clean(unit.get("final_template_column"))
    client = clean(unit.get("client_name"))
    note = clean(unit.get("final_note") or unit.get("expense_note") or unit.get("source_note"))
    errors: list[str] = []
    if source_category != "mobile":
        if final_column == "mobile":
            errors.append("non-mobile expense cannot use the mobile amount column")
        if client == MOBILE_CLIENT:
            errors.append("non-mobile expense cannot use Client = 通讯费")
        if MOBILE_CLIENT in note:
            errors.append("non-mobile expense cannot use a 通讯费 note")
    else:
        if final_column != "mobile":
            errors.append("mobile expense must use the mobile amount column")
        if clean(unit.get("client_charge_code")).upper() != ADMIN_CODE:
            errors.append(f"mobile expense must be assigned to {ADMIN_CODE}")
    project_expense_categories = {"hotel", "meal", "taxi", "travel"}
    if source_category in project_expense_categories or final_column in project_expense_categories:
        if clean(unit.get("client_charge_code")).upper() == ADMIN_CODE:
            errors.append(f"project expenses cannot be assigned to {ADMIN_CODE}")
    return errors


def formal_meal_column(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) != "meal":
        return clean(unit.get("final_template_column"))
    city = clean(unit.get("city"))
    if city and "上海" in city:
        return "meal"
    if city:
        return "travel"
    current = clean(unit.get("final_template_column"))
    # A stale non-meal column (e.g. "other" left over from reclassification)
    # must not survive: it would drop the row into the wrong workbook column
    # and hide it from the daily meal cap policy detection.
    return current if current in {"meal", "travel"} else "meal"


def is_shanghai_city(value: Any) -> bool:
    text = clean(value).lower()
    return "上海" in text or "shanghai" in text


def is_ride_unit(unit: dict[str, Any]) -> bool:
    return bool(
        clean(unit.get("origin"))
        or clean(unit.get("destination"))
        or clean(unit.get("source_item_id"))
        or clean(unit.get("document_subtype")) in {"didi_trip_report", "gaode_trip_report"}
    )


def formal_taxi_column(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) not in {"taxi", "travel"} or not is_ride_unit(unit):
        return clean(unit.get("final_template_column"))
    city = clean(unit.get("city"))
    if city and not is_shanghai_city(city):
        return "travel"
    if city:
        return "taxi"
    return clean(unit.get("final_template_column")) or ("taxi" if clean(unit.get("source_category")) == "taxi" else "travel")


def formal_amount_column(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) == "meal":
        return formal_meal_column(unit)
    if is_ride_unit(unit):
        return formal_taxi_column(unit)
    return clean(unit.get("final_template_column"))


def contains_place_type_placeholder(note: Any) -> bool:
    text = clean(note)
    return "出发地类型" in text or "目的地类型" in text


def contains_hotel_placeholder(note: Any) -> bool:
    text = clean(note)
    return any(token in text for token in ["X晚", "入住日", "离店日"])


def taxi_template_note(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) not in {"taxi", "travel"} or not clean(unit.get("origin")):
        return ""
    origin_type = clean(unit.get("origin_place_type"))
    dest_type = clean(unit.get("destination_place_type"))
    if not origin_type or not dest_type:
        return ""
    suffix = "（加班）" if clean(unit.get("business_reason")) == "overtime" else ""
    return f"打车（{origin_type}-{dest_type}）{suffix}"


def meal_template_note(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) != "meal":
        return ""
    note = clean(unit.get("final_note") or unit.get("expense_note") or unit.get("source_note"))
    allowed_prefixes = (C["trip_meal"], "出差餐费（高铁站/机场）", C["overtime_meal"])
    if note.startswith(allowed_prefixes):
        return note
    context = clean(unit.get("meal_context"))
    if context == "overtime":
        return C["overtime_meal"]
    if context == "station_airport":
        return "出差餐费（高铁站/机场）"
    return C["trip_meal"]


def classify_meal_policy(value: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    """Classify meal policy by substantive purpose, never by workbook form fields."""
    if clean(value.get("source_category")) != "meal":
        return None, None

    # Event-declared, one-off daily cap wins over the generic two-tier policy.
    # It is stamped by annotate_event_meal_standards() from the meal's
    # project_context meal_standards; never inferred from city/amount column.
    event_cap = clean(value.get("event_meal_cap"))
    if event_cap:
        context_id = clean(value.get("event_meal_context_id"))
        label = clean(value.get("event_meal_label")) or "事件餐标"
        direction = clean(value.get("event_meal_direction")) or "unknown"
        return {
            "policy": f"event_declared:{context_id}" if context_id else "event_declared",
            "policy_name": label,
            "cap": Decimal(money(event_cap)),
            "basis": [
                f"event-declared daily cap {money(event_cap)} from "
                f"{context_id or 'project_context'} ({label}); user-declared via "
                f"project_context.meal_standards; direction={direction}"
            ],
        }, None

    note = clean(value.get("note") or value.get("final_note"))
    context = clean(value.get("meal_context"))
    trip_signals: list[str] = []
    overtime_signals: list[str] = []
    if note.startswith(C["trip_meal"]):
        trip_signals.append("final_note starts with 出差餐费")
    if note.startswith(C["overtime_meal"]):
        overtime_signals.append("final_note starts with 加班餐费")
    if context in TRIP_MEAL_CONTEXTS:
        trip_signals.append(f"meal_context={context}")
    if context in OVERTIME_MEAL_CONTEXTS:
        overtime_signals.append(f"meal_context={context}")

    if trip_signals and overtime_signals:
        return None, (
            "meal policy signals conflict: " + ", ".join(trip_signals + overtime_signals)
            + "; city/amount column cannot resolve this conflict"
        )
    if trip_signals:
        return {
            "policy": "business_trip_meal",
            "policy_name": C["business_trip_meal_policy"],
            "cap": BUSINESS_TRIP_MEAL_DAILY_CAP,
            "basis": trip_signals,
        }, None
    if overtime_signals:
        return {
            "policy": "local_overtime_meal",
            "policy_name": C["local_overtime_meal_policy"],
            "cap": LOCAL_OVERTIME_MEAL_DAILY_CAP,
            "basis": overtime_signals,
        }, None
    return None, (
        "meal policy is missing: explicitly set final_note to 出差餐费/出差餐费（高铁站/机场）/加班餐费 "
        "or set meal_context to business_trip/station_airport/overtime; do not infer it from Shanghai, "
        "amount_column, or Expense Nature"
    )


def hotel_template_note(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) != "hotel":
        return ""
    note = clean(unit.get("final_note") or unit.get("expense_note") or unit.get("source_note"))
    if note.startswith("出差酒店") and "晚" in note and not contains_hotel_placeholder(note):
        return note
    nights = clean(unit.get("hotel_nights")) or "X"
    checkin = clean(unit.get("check_in_date")) or "入住日"
    checkout = clean(unit.get("check_out_date")) or "离店日"
    return f"出差酒店（{nights}晚，{checkin}-{checkout}）"


def mobile_template_note(unit: dict[str, Any]) -> str:
    if clean(unit.get("source_category")) != "mobile":
        return ""
    note = clean(unit.get("final_note") or unit.get("expense_note") or unit.get("source_note"))
    if "通讯费" in note:
        return note
    period = clean(unit.get("billing_period"))
    if period and len(period) >= 6 and period[:4].isdigit() and period[4:6].isdigit():
        return f"{int(period[4:6])}月通讯费"
    return "X月通讯费"


def normalized_note_base(unit: dict[str, Any]) -> str:
    note = clean(unit.get("final_note") or unit.get("expense_note") or unit.get("source_note"))
    template_note = ticket_note(unit)
    source_note = clean(unit.get("source_note") or unit.get("expense_note"))
    if template_note and (not note or note == source_note or contains_raw_ticket_evidence(note)):
        return template_note
    taxi_note = taxi_template_note(unit)
    if taxi_note and (not note or note == source_note or "->" in note or contains_place_type_placeholder(note)):
        return taxi_note
    for note_builder in (meal_template_note, hotel_template_note, mobile_template_note):
        built = note_builder(unit)
        if built:
            return built
    return note


def stage3_note_errors(unit: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if contains_place_type_placeholder(unit.get("final_note")):
        errors.append("taxi note still contains place-type placeholders")
    if (
        clean(unit.get("status")) in {"confirmed", "fixed"}
        and clean(unit.get("source_category")) == "hotel"
        and contains_hotel_placeholder(unit.get("final_note"))
        and contains_hotel_placeholder(normalized_note_base(unit))
    ):
        errors.append("hotel final_note still contains night/date placeholders")
    if (
        clean(unit.get("status")) in {"confirmed", "fixed"}
        and clean(unit.get("source_category")) in {"taxi", "travel"}
        and clean(unit.get("origin"))
        and (not clean(unit.get("origin_place_type")) or not clean(unit.get("destination_place_type")))
    ):
        errors.append("confirmed taxi/travel item requires origin_place_type and destination_place_type")
    if clean(unit.get("status")) in {"confirmed", "fixed"} and (is_rail_ticket(unit) or is_flight_ticket(unit)):
        if not ticket_note(unit):
            errors.append("rail/flight item requires route evidence for the final note template")
        elif contains_raw_ticket_evidence(unit.get("final_note")):
            errors.append("rail/flight final_note must use the finance template, not raw ticket evidence")
    return errors


def stage3_rule_errors(unit: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    category = clean(unit.get("source_category"))
    if category not in ALLOWED_SOURCE_CATEGORIES:
        errors.append(f"unsupported source_category: {category or '<blank>'}")

    visible_column = formal_amount_column(unit)
    if visible_column not in AMOUNT_COLUMNS:
        errors.append(f"invalid final amount column: {visible_column or '<blank>'}")
    if (is_rail_ticket(unit) or is_flight_ticket(unit)) and visible_column != "travel":
        errors.append("rail/flight items must use the travel amount column")
    if category == "hotel" and visible_column != "hotel":
        errors.append("hotel expenses must use the hotel amount column")
    if category == "hotel":
        missing_hotel_fields = [
            field for field in ["hotel_nights", "check_in_date", "check_out_date"]
            if not clean(unit.get(field))
        ]
        if missing_hotel_fields:
            errors.append(
                "hotel expenses require hotel_nights, check_in_date, and check_out_date before writing the workbook"
            )
        if contains_hotel_placeholder(normalized_note_base(unit)):
            errors.append("hotel final note must use actual nights/check-in/check-out, not X晚/入住日/离店日 placeholders")
    if category == "mobile" and visible_column != "mobile":
        errors.append("mobile expenses must use the mobile amount column")
    if category == "meal":
        _policy, policy_error = classify_meal_policy(unit)
        if policy_error:
            errors.append(policy_error)

    for field in ("amount", "invoice_amount", "reimbursable_amount"):
        value = unit.get(field)
        if value_utils.is_blank(value):
            continue
        try:
            value_utils.parse_finite_decimal(value, field=field)
        except ValueError as exc:
            errors.append(str(exc))

    # A doc/template placeholder such as <ADMIN_CODE>/<BD_CODE>/CORP-<FY>-ADMIN must
    # never reach the workbook. Angle brackets never appear in a real fiscal-year
    # code, so treat any as an unresolved placeholder and block generation.
    code_text = clean(unit.get("client_charge_code"))
    if "<" in code_text or ">" in code_text:
        errors.append(
            f"client_charge_code still contains a template placeholder ({code_text!r}); "
            "use the current fiscal-year BD/ADMIN code from assets/special-code-definitions.json "
            "(roll it with special_codes.py set-year), never a <...> placeholder"
        )

    if not normalized_note_base(unit) and category != "other":
        errors.append("final note cannot be derived from the confirmed allocation")
    if parse_date(unit.get("expense_date")) is None:
        errors.append("expense_date must be a valid YYYY-MM-DD date before writing the workbook")

    return errors


money = value_utils.format_money


def invoice_amount(unit: dict[str, Any]) -> str:
    return money(value_utils.first_nonblank(unit.get("invoice_amount"), unit.get("amount")))


def reimbursable_amount(unit: dict[str, Any]) -> str:
    return money(value_utils.first_nonblank(unit.get("reimbursable_amount"), unit.get("amount")))


def date_yyyymmdd(value: str) -> str:
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", value or "")
    if not match:
        compact = re.sub(r"\D", "", value or "")
        return compact[:8]
    return f"{int(match.group(1)):04d}{int(match.group(2)):02d}{int(match.group(3)):02d}"


def parse_date(value: Any) -> date | None:
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", clean(value))
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
    except ValueError:
        return None


def date_in_context(unit_date: Any, ctx: dict[str, Any]) -> bool:
    d = parse_date(unit_date)
    start = parse_date(ctx.get("date_start", ""))
    end = parse_date(ctx.get("date_end", ""))
    if not d or not start or not end:
        return False
    buffer = int(ctx.get("travel_buffer_days") or 0)
    return start - timedelta(days=buffer) <= d <= end + timedelta(days=buffer)


def route_endpoints(unit: dict[str, Any]) -> tuple[str, str]:
    route = (
        route_from_text(unit.get("route"))
        or route_from_text(unit.get("source_note"))
        or route_from_text(unit.get("expense_note"))
        or route_from_text(unit.get("final_note"))
    )
    if not route:
        return "", ""
    parts = route.split("-", 1)
    if len(parts) == 2:
        return clean(parts[0]), clean(parts[1])
    return "", ""


def rail_station_key(value: Any) -> str:
    text = clean(value).lower().replace(" ", "")
    for suffix in ["火车站", "高铁站", "铁路站", "站"]:
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    return text


def rail_station_city_key(value: Any) -> str:
    text = rail_station_key(value)
    if len(text) >= 3 and text[-1:] in {"东", "西", "南", "北"}:
        text = text[:-1]
    return text.replace("市", "")


def rail_stations_connect(destination: Any, origin: Any) -> bool:
    destination_station = rail_station_key(destination)
    origin_station = rail_station_key(origin)
    if not destination_station or not origin_station:
        return False
    return (
        destination_station == origin_station
        or rail_station_city_key(destination_station) == rail_station_city_key(origin_station)
    )


def rail_chain_groups(units: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for unit in units:
        chain_id = clean(unit.get("journey_chain_id"))
        if chain_id and clean(unit.get("status")) not in {"dropped", "excluded", "non_reimbursable"}:
            groups[chain_id].append(unit)
    for chain_units in groups.values():
        chain_units.sort(key=lambda unit: integer_or_zero(unit.get("journey_chain_position")))
    return dict(groups)


def rail_chain_route(chain_units: list[dict[str, Any]]) -> str:
    if not chain_units:
        return ""
    first_origin, _ = route_endpoints(chain_units[0])
    destinations = [route_endpoints(unit)[1] for unit in chain_units]
    if not first_origin or any(not destination for destination in destinations):
        return ""
    return " -> ".join([first_origin, *destinations])


def integer_or_zero(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def rail_chain_ready_errors(units: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    for chain_id, chain_units in rail_chain_groups(units).items():
        if len(chain_units) < 2:
            errors.append(f"{chain_id} has fewer than two active rail legs; rerun Stage 2 to rebuild journey chains")
            continue
        if any(unit.get("journey_chain_needs_confirmation") for unit in chain_units):
            errors.append(f"{chain_id} still needs one whole-journey project decision")

        assignments = {
            (
                clean(unit.get("project_context_id")),
                clean(unit.get("client_name")),
                clean(unit.get("client_charge_code")),
            )
            for unit in chain_units
        }
        if len(assignments) != 1:
            numbers = ", ".join(str(unit.get("user_no") or unit.get("unit_id")) for unit in chain_units)
            errors.append(
                f"{chain_id} items {numbers} are one connected rail journey but have different project assignments; "
                "update all legs together or rerun Stage 2 if the chain itself is wrong"
            )

        expected_positions = list(range(1, len(chain_units) + 1))
        actual_positions = [integer_or_zero(unit.get("journey_chain_position")) for unit in chain_units]
        if actual_positions != expected_positions:
            errors.append(f"{chain_id} has stale or duplicate leg positions; rerun Stage 2")

        declared_lengths = {
            integer_or_zero(unit.get("journey_chain_length"))
            for unit in chain_units
        }
        if declared_lengths != {len(chain_units)}:
            errors.append(
                f"{chain_id} has stale chain length metadata, usually because a leg was dropped; rerun Stage 2"
            )

        active_unit_ids = [clean(unit.get("unit_id")) for unit in chain_units]
        declared_member_lists = {
            tuple(clean(unit_id) for unit_id in (unit.get("journey_chain_unit_ids") or []))
            for unit in chain_units
        }
        if declared_member_lists != {tuple(active_unit_ids)}:
            errors.append(
                f"{chain_id} has stale chain member metadata, usually because a leg was dropped; rerun Stage 2"
            )

        current_route = rail_chain_route(chain_units)
        declared_routes = {clean(unit.get("journey_chain_route")) for unit in chain_units}
        if not current_route or declared_routes != {current_route}:
            errors.append(f"{chain_id} has stale whole-journey route metadata; rerun Stage 2")

        for first, second in zip(chain_units, chain_units[1:]):
            _, first_destination = route_endpoints(first)
            second_origin, _ = route_endpoints(second)
            if not rail_stations_connect(first_destination, second_origin):
                errors.append(
                    f"{chain_id} is no longer continuous at {first_destination or '?'} -> {second_origin or '?'}; "
                    "a route correction invalidated the chain, so rerun Stage 2"
                )
    return errors


def rail_chain_summaries(units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for chain_id, chain_units in rail_chain_groups(units).items():
        first = chain_units[0]
        summaries.append({
            "journey_chain_id": chain_id,
            "route": clean(first.get("journey_chain_route")),
            "unit_ids": [clean(unit.get("unit_id")) for unit in chain_units],
            "user_nos": [unit.get("user_no", "") for unit in chain_units],
            "project_context_id": clean(first.get("project_context_id")),
            "client_name": clean(first.get("client_name")),
            "client_charge_code": clean(first.get("client_charge_code")),
            "confidence": clean(first.get("journey_chain_confidence")),
            "assignment_rule": clean(first.get("journey_chain_assignment_rule")),
            "needs_confirmation": any(unit.get("journey_chain_needs_confirmation") for unit in chain_units),
        })
    return summaries


def assignment_matches_context(unit: dict[str, Any], ctx: dict[str, Any]) -> bool:
    if clean(unit.get("project_context_id")) and clean(unit.get("project_context_id")) == clean(ctx.get("context_id")):
        return True
    return (
        clean(unit.get("client_name")) == clean(ctx.get("client_name"))
        and clean(unit.get("client_charge_code")) == clean(ctx.get("client_charge_code"))
    )


def travel_destination_context(unit: dict[str, Any], contexts: list[dict[str, Any]]) -> dict[str, Any] | None:
    if clean(unit.get("source_category")) != "travel" and clean(unit.get("document_subtype")) != "railway_e_ticket":
        return None
    _, destination = route_endpoints(unit)
    if not destination:
        return None
    candidates = [
        ctx for ctx in contexts
        if date_in_context(unit.get("expense_date", ""), ctx)
        and clean(ctx.get("city"))
        and clean(ctx.get("city")) in destination
        # Admin is not a project you travel to; never treat it as a destination
        # match (consistent with allocate_expenses.non_admin_contexts).
        and clean(ctx.get("client_charge_code")).upper() != ADMIN_CODE
    ]
    context_ids = {clean(ctx.get("context_id")) for ctx in candidates}
    if len(context_ids) == 1:
        return candidates[0]
    return None


def expense_hint_reconciliation_errors(allocation: dict[str, Any]) -> list[str]:
    contexts_have_hints = any(
        context.get(field)
        for context in allocation.get("project_contexts", [])
        for field in ("meal_hints", "expense_hints")
    )
    if "expense_hint_reconciliation" not in allocation:
        if contexts_have_hints:
            return [
                "User expense hints exist but the reverse reconciliation ledger is missing; rerun Stage 2."
            ]
        return []
    records = allocation.get("expense_hint_reconciliation")
    if not isinstance(records, list):
        return ["User expense hint reconciliation must be a list; rerun Stage 2."]

    units_by_id = {
        clean(unit.get("unit_id")): unit
        for unit in allocation.get("allocation_units", [])
        if clean(unit.get("unit_id"))
    }
    errors: list[str] = []
    for record in records:
        hint_id = clean(record.get("hint_id")) or "<unknown hint>"
        resolution_status = clean(record.get("resolution_status"))
        match_status = clean(record.get("match_status"))
        if resolution_status not in {"not_required", "open", "pending_evidence", "resolved"}:
            errors.append(
                f"User expense record {hint_id} has invalid resolution status {resolution_status!r}; rerun Stage 2."
            )
            continue
        if resolution_status == "open":
            errors.append(
                f"User expense record {hint_id} ({record.get('summary', '')}) has no unique invoice match or explicit resolution."
            )
            continue
        if resolution_status == "pending_evidence":
            errors.append(
                f"User expense record {hint_id} ({record.get('summary', '')}) is still waiting for an invoice; "
                "supply the evidence or explicitly mark the record not reimbursed."
            )
            continue
        if resolution_status == "resolved":
            action = clean(record.get("resolution_action"))
            if action not in {"matched_existing", "covered_by_invoice", "not_reimbursed"}:
                errors.append(
                    f"User expense record {hint_id} has no valid structured resolution action; resolve it again in Stage 2."
                )
            if not clean(record.get("resolution_answer")):
                errors.append(f"User expense record {hint_id} was closed without an explicit resolution answer.")
        if resolution_status == "not_required" and match_status != "matched":
            errors.append(
                f"User expense record {hint_id} skips explicit resolution without a unique matched unit."
            )
        if match_status in {"matched", "covered"}:
            matched_ids = [clean(value) for value in record.get("matched_unit_ids", []) if clean(value)]
            active_matches = [
                unit_id for unit_id in matched_ids
                if unit_id in units_by_id
                and clean(units_by_id[unit_id].get("status")) not in {"dropped", "excluded", "non_reimbursable"}
            ]
            if not matched_ids or not active_matches:
                errors.append(
                    f"User expense record {hint_id} points only to a missing/dropped expense unit; resolve it again."
                )
    return errors


def require_ready(allocation: dict[str, Any], allow_unconfirmed: bool) -> list[str]:
    errors: list[str] = []
    contexts = allocation.get("project_contexts", [])
    errors.extend(rail_chain_ready_errors(allocation.get("allocation_units", [])))
    hint_errors = expense_hint_reconciliation_errors(allocation)
    if hint_errors and not allow_unconfirmed:
        errors.extend(hint_errors)
    open_questions = [q for q in allocation.get("questions", []) if q.get("status", "open") == "open"]
    if open_questions and not allow_unconfirmed:
        errors.append(f"{len(open_questions)} open allocation question(s) remain.")
    for unit in allocation.get("allocation_units", []):
        status = unit.get("status", "")
        if status in {"dropped", "excluded", "non_reimbursable"}:
            continue
        if not allow_unconfirmed and status not in {"confirmed", "fixed"}:
            errors.append(f"{unit.get('unit_id')} is not confirmed or fixed.")
        if "餐费" in clean(unit.get("final_note")) and clean(unit.get("source_category")) != "meal":
            errors.append(
                f"{unit.get('unit_id')} has a meal-style final_note but source_category "
                f"{clean(unit.get('source_category'))!r} — it would escape the daily meal cap check. "
                "Fix source_category (or the note) via the answers updater before writing the workbook."
            )
        if not allow_unconfirmed and unit.get("date_required"):
            errors.append(f"{unit.get('unit_id')} still requires a user-confirmed expense date.")
        if (
            not allow_unconfirmed
            and unit.get("date_is_provisional")
            and unit.get("source_category") != "other"
        ):
            errors.append(f"{unit.get('unit_id')} has a provisional date but is not an other expense.")
        for accounting_error in mobile_accounting_errors(unit):
            errors.append(f"{unit.get('unit_id')} accounting conflict: {accounting_error}.")
        for note_error in stage3_note_errors(unit):
            errors.append(f"{unit.get('unit_id')} note conflict: {note_error}.")
        for rule_error in stage3_rule_errors(unit):
            errors.append(f"{unit.get('unit_id')} stage-3 rule conflict: {rule_error}.")
        destination_ctx = None if clean(unit.get("journey_chain_id")) else travel_destination_context(unit, contexts)
        if destination_ctx and not assignment_matches_context(unit, destination_ctx):
            errors.append(
                f"{unit.get('unit_id')} travel route destination points to "
                f"{destination_ctx.get('client_name', '')}/{destination_ctx.get('client_charge_code', '')}; "
                "travel between project cities should belong to the project being traveled to."
            )
        for field in ["client_name", "client_charge_code", "final_template_column", "amount", "expense_date"]:
            if field == "client_name" and clean(unit.get("client_charge_code")).upper() == ADMIN_CODE:
                continue
            if not unit.get(field):
                errors.append(f"{unit.get('unit_id')} missing {field}.")
    return errors


def print_stage3_preflight_summary(allocation: dict[str, Any], errors: list[str]) -> None:
    included = [
        unit for unit in allocation.get("allocation_units", [])
        if unit.get("status") not in {"dropped", "excluded", "non_reimbursable"}
    ]
    open_questions = [q for q in allocation.get("questions", []) if q.get("status", "open") == "open"]
    print("\nSTAGE 3 PREFLIGHT CHECK TO SHOW IN CHAT")
    print(f"units={len(included)} open_questions={len(open_questions)} blocking_errors={len(errors)}")
    if errors:
        print("ACTION REQUIRED: Do not write the initial reimbursement workbook until these allocation issues are fixed.")
        for error in errors[:20]:
            print(f"- {error}")
        if len(errors) > 20:
            print(f"- ... {len(errors) - 20} more error(s)")
        return
    print("OK: allocation is structurally ready for initial workbook generation.")


def included_units(allocation: dict[str, Any]) -> list[dict[str, Any]]:
    out = []
    for unit in allocation.get("allocation_units", []):
        if unit.get("status") in {"dropped", "excluded", "non_reimbursable"}:
            continue
        if Decimal(reimbursable_amount(unit)) == 0:
            continue
        out.append(dict(unit))
    return out


# Neutral fallback label for a supporting document whose type the user left unset.
SUPPORT_DOC_DEFAULT_TYPE = "支持文档"  # 支持文档


def _resolve_path(value: str) -> str:
    value = clean(value)
    if not value:
        return ""
    try:
        return str(Path(value).resolve())
    except OSError:
        return value


def _unit_invoice_doc_ids(units: list[dict[str, Any]]) -> set[str]:
    """Document ids that a supporting document may legitimately name as its invoice."""
    ids: set[str] = set()
    for unit in units:
        for field in ("supporting_invoice_document_id", "source_document_id"):
            value = clean(unit.get(field))
            if value:
                ids.add(value)
    return ids


def _substitute_approval_sources(units: list[dict[str, Any]]) -> set[str]:
    """Resolved paths already packaged via the substitute-invoice approval_file path."""
    out: set[str] = set()
    for unit in units:
        if unit.get("is_substitute_invoice"):
            resolved = _resolve_path(unit.get("approval_file", ""))
            if resolved:
                out.add(resolved)
    return out


def collect_support_documents(
    extraction: dict[str, Any], units: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Split standalone supporting documents into mounted vs. orphan.

    A supporting document is *mounted* when it names, via ``supports_document_id``,
    an invoice that survives into an included expense unit — it will be packaged
    under that invoice's proof number. A substitute-invoice approval screenshot is
    already carried through the substitute unit's ``approval_file`` and is skipped
    here to avoid double packaging. Everything else is an *orphan*: evidence the
    user kept but did not tie to any invoice, which hard-blocks Stage 3.
    """
    valid_invoice_ids = _unit_invoice_doc_ids(units)
    substitute_sources = _substitute_approval_sources(units)
    mounted: list[dict[str, Any]] = []
    orphans: list[dict[str, Any]] = []
    for doc in extraction.get("documents", []):
        if doc.get("excluded_by_user"):
            continue
        if doc.get("document_role") != "supporting_document":
            continue
        source = clean(doc.get("source_file"))
        if source and _resolve_path(source) in substitute_sources:
            continue
        supports = clean(doc.get("supports_document_id"))
        if supports and supports in valid_invoice_ids:
            mounted.append({
                "document_id": doc.get("document_id"),
                "source_file": source,
                "support_type": clean(doc.get("support_type")) or SUPPORT_DOC_DEFAULT_TYPE,
                "supports_document_id": supports,
            })
        else:
            orphans.append(doc)
    return mounted, orphans


def attach_support_documents_to_groups(
    proof_groups: list[dict[str, Any]], mounted: list[dict[str, Any]]
) -> None:
    """Add each mounted support document to the proof group of the invoice it backs."""
    group_by_invoice: dict[str, dict[str, Any]] = {}
    for group in proof_groups:
        for doc_id in group.get("source_document_ids", []):
            group_by_invoice.setdefault(doc_id, group)
    for item in mounted:
        group = group_by_invoice.get(item["supports_document_id"])
        if group is None:
            continue  # unreachable once preflight has passed; guard defensively
        group.setdefault("support_documents", []).append({
            "document_id": item["document_id"],
            "source_file": item["source_file"],
            "support_type": item["support_type"],
        })


def proof_type(unit: dict[str, Any]) -> str:
    subtype = unit.get("document_subtype", "")
    source = unit.get("source_category", "")
    seller = clean(unit.get("seller_name", ""))
    source_doc = clean(unit.get("source_note", ""))
    if subtype == "railway_e_ticket":
        return "rail"
    if source == "meal":
        return "meal"
    if is_flight_ticket(unit) or source == "flight":
        return "flight"
    if is_rail_ticket(unit) or source == "rail":
        return "rail"
    if source == "hotel":
        return "hotel"
    if subtype == "gaode_trip_report" or "高德" in seller or "高德" in source_doc:
        return "gaode"
    if unit.get("source_item_id") or "滴滴" in seller or "Didi" in source_doc:
        return "taxi_didi"
    if source == "taxi":
        return "taxi"
    if source == "mobile":
        return "mobile"
    if source == "travel":
        return "travel"
    return source or "other"


def proof_key(unit: dict[str, Any]) -> str:
    ptype = proof_type(unit)
    if ptype in {"taxi_didi", "gaode"}:
        return unit.get("supporting_invoice_document_id") or unit.get("supporting_schedule_document_id") or unit.get("source_document_id")
    return unit.get("supporting_invoice_document_id") or unit.get("source_document_id")


def assign_proof_numbers(units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for unit in units:
        ptype = proof_type(unit)
        key = proof_key(unit) or unit["unit_id"]
        gkey = (ptype, key)
        group = groups.setdefault(gkey, {
            "proof_group_id": f"PROOF-{len(groups)+1:03d}",
            "proof_type": ptype,
            "proof_key": key,
            "source_document_ids": [],
            "support_document_ids": [],
            "source_item_ids": [],
            "source_invoice_no": unit.get("invoice_no", ""),
            "amount_total": Decimal("0.00"),
            "min_date": unit.get("expense_date", ""),
            "units": [],
        })
        group["units"].append(unit["unit_id"])
        group["amount_total"] += Decimal(reimbursable_amount(unit))
        group["min_date"] = min([d for d in [group["min_date"], unit.get("expense_date", "")] if d] or [""])
        for field, target in [
            ("supporting_invoice_document_id", "source_document_ids"),
            ("source_document_id", "source_document_ids"),
            ("supporting_schedule_document_id", "support_document_ids"),
        ]:
            value = unit.get(field)
            if value and value not in group[target]:
                group[target].append(value)
        if unit.get("source_item_id") and unit["source_item_id"] not in group["source_item_ids"]:
            group["source_item_ids"].append(unit["source_item_id"])

    ordered = sorted(
        groups.values(),
        key=lambda g: (PROOF_ORDER.get(g["proof_type"], 99), g.get("min_date", ""), g["proof_key"]),
    )
    for idx, group in enumerate(ordered, start=1):
        group["proof_no"] = idx
        group["amount_total"] = money(group["amount_total"])
    proof_by_unit = {}
    for group in ordered:
        for unit_id in group["units"]:
            proof_by_unit[unit_id] = group["proof_no"]
    for unit in units:
        unit["proof_no"] = proof_by_unit[unit["unit_id"]]
    return ordered


def final_note(unit: dict[str, Any]) -> str:
    note = normalized_note_base(unit)
    if unit.get("is_substitute_invoice") and C["substitute"] not in note:
        note += C["substitute"]
    invoice = Decimal(invoice_amount(unit))
    reimbursable = Decimal(reimbursable_amount(unit))
    if reimbursable != invoice and C["invoice_amount"] not in note:
        note += f"\uff08{C['invoice_amount']}{money(invoice)}/{C['reimbursable_amount']}{money(reimbursable)}\uff09"
    return note


def expense_nature(unit: dict[str, Any]) -> str:
    city = clean(unit.get("city"))
    amount_column = formal_amount_column(unit)
    if amount_column == "mobile":
        return C["local"]
    if amount_column in {"hotel", "travel"}:
        return C["trip"]
    if city:
        return C["local"] if "\u4e0a\u6d77" in city else C["trip"]
    value = clean(unit.get("expenses_nature"))
    if value:
        return value
    return C["local"]


def make_rows(units: list[dict[str, Any]], requester: str) -> list[dict[str, Any]]:
    rows = []
    for unit in units:
        amount_col = formal_amount_column(unit)
        amount_col = amount_col or "other"
        if amount_col not in AMOUNT_COLUMNS:
            amount_col = "other"
        rows.append({
            "date": date_yyyymmdd(unit.get("expense_date", "")),
            "requester": requester,
            "client": normalized_client_name(unit),
            "client_charge_code": unit.get("client_charge_code", ""),
            "project_context_id": clean(unit.get("project_context_id")),
            "expenses_nature": expense_nature(unit),
            "note": final_note(unit),
            "amount_column": amount_col,
            "amount": reimbursable_amount(unit),
            "invoice_amount": invoice_amount(unit),
            "reimbursable_amount": reimbursable_amount(unit),
            "proof_no": unit.get("proof_no"),
            "user_no": unit.get("user_no", ""),
            "source_unit_id": unit.get("unit_id"),
            "source_document_id": unit.get("source_document_id", ""),
            "source_item_id": unit.get("source_item_id", ""),
            "source_category": unit.get("source_category", ""),
            "source_filename": unit.get("source_filename", ""),
            "supporting_invoice_filename": unit.get("supporting_invoice_filename", ""),
            "issue_date": unit.get("issue_date", ""),
            "date_source": unit.get("date_source", ""),
            "date_is_provisional": unit.get("date_is_provisional", False),
            "date_required": unit.get("date_required", False),
            "seller_name": unit.get("seller_name", ""),
            "city": unit.get("city", ""),
            "attendees": unit.get("attendees", ""),
            "meal_context": unit.get("meal_context", ""),
            "train_no": unit.get("train_no", ""),
            "origin_station": unit.get("origin_station", ""),
            "destination_station": unit.get("destination_station", ""),
            "rail_departure_time": unit.get("rail_departure_time", ""),
            "rail_departure_datetime": unit.get("rail_departure_datetime", ""),
            "journey_chain_id": unit.get("journey_chain_id", ""),
            "journey_chain_route": unit.get("journey_chain_route", ""),
            "journey_chain_position": unit.get("journey_chain_position", ""),
            "journey_chain_length": unit.get("journey_chain_length", ""),
            "journey_chain_unit_ids": unit.get("journey_chain_unit_ids", []),
            "journey_chain_confidence": unit.get("journey_chain_confidence", ""),
            "journey_chain_assignment_rule": unit.get("journey_chain_assignment_rule", ""),
            "journey_chain_match_reason": unit.get("journey_chain_match_reason", ""),
            "journey_chain_project_context_id": unit.get("journey_chain_project_context_id", ""),
            "journey_chain_needs_confirmation": bool(unit.get("journey_chain_needs_confirmation")),
            "hotel_city": unit.get("hotel_city") or unit.get("city", ""),
            "hotel_city_tier": unit.get("hotel_city_tier", ""),
            "hotel_nights": unit.get("hotel_nights", ""),
            "check_in_date": unit.get("check_in_date", ""),
            "check_out_date": unit.get("check_out_date", ""),
            "shared_room": unit.get("shared_room", False),
            "room_shared_with": unit.get("room_shared_with", ""),
            "room_share_note": unit.get("room_share_note", ""),
            "is_substitute_invoice": bool(unit.get("is_substitute_invoice")),
            "substitute_for": unit.get("substitute_for", ""),
            "approval_required": unit.get("approval_required", ""),
            "approval_file": unit.get("approval_file", ""),
            "approval_file_status": unit.get("approval_file_status", ""),
            "manual_correction": bool(unit.get("manual_correction")),
            "correction_note": unit.get("correction_note", ""),
            "corrected_fields": unit.get("corrected_fields", []),
            "row_order_type": proof_type(unit),
            "expense_date": unit.get("expense_date", ""),
        })
    return rows


def meal_cap_policy(row: dict[str, Any]) -> dict[str, Any] | None:
    policy, _error = classify_meal_policy(row)
    return policy


def _context_meal_standard(
    contexts: list[dict[str, Any]], context_id: str, normalized_date: str
) -> dict[str, Any] | None:
    """One-off daily meal cap the user declared for this context on this date.

    Matches strictly on context_id + date so a same-date expense in another
    context is never captured by an event standard.
    """
    if not context_id or not normalized_date:
        return None
    for ctx in contexts:
        if not isinstance(ctx, dict) or clean(ctx.get("context_id")) != context_id:
            continue
        for standard in ctx.get("meal_standards", []) or []:
            if not isinstance(standard, dict):
                continue
            if date_yyyymmdd(clean(standard.get("date"))) != normalized_date:
                continue
            cap = clean(standard.get("daily_cap"))
            if not cap:
                return None
            return {
                "context_id": context_id,
                "cap": cap,
                "label": clean(standard.get("label")) or "事件餐标",
                "basis": clean(standard.get("basis")) or "用户声明",
            }
        return None
    return None


def annotate_event_meal_standards(
    rows: list[dict[str, Any]], contexts: list[dict[str, Any]]
) -> None:
    """Stamp user-declared event daily caps onto meal rows before classification.

    Records the effective cap plus its direction against the generic policy that
    would otherwise apply, so the deterministic writer and the Gate Challenger
    both see an auditable provenance for the override.
    """
    contexts = contexts or []
    for row in rows:
        for field in ("event_meal_cap", "event_meal_label", "event_meal_context_id", "event_meal_direction"):
            row.pop(field, None)
        if clean(row.get("source_category")) != "meal":
            continue
        context_id = clean(row.get("project_context_id"))
        normalized_date = clean(row.get("date")) or date_yyyymmdd(row.get("expense_date", ""))
        standard = _context_meal_standard(contexts, context_id, normalized_date)
        if not standard:
            continue
        # Compute direction against the generic cap that would apply without the
        # override. classify_meal_policy still returns the generic tier here
        # because event_meal_cap has not been stamped on the row yet.
        generic_policy, _error = classify_meal_policy(row)
        if generic_policy is None:
            direction = "no_generic_baseline"
        else:
            event_cap = Decimal(money(standard["cap"]))
            generic_cap = generic_policy["cap"]
            if event_cap > generic_cap:
                direction = "exceeds_generic"
            elif event_cap < generic_cap:
                direction = "conservative"
            else:
                direction = "equal"
        row["event_meal_cap"] = money(standard["cap"])
        row["event_meal_label"] = standard["label"]
        row["event_meal_context_id"] = standard["context_id"]
        row["event_meal_direction"] = direction


def annotate_meal_policies(rows: list[dict[str, Any]]) -> None:
    for row in rows:
        if clean(row.get("source_category")) != "meal":
            continue
        policy, error = classify_meal_policy(row)
        if error or not policy:
            row["meal_cap_policy"] = "unclassified"
            row["meal_daily_cap"] = ""
            row["meal_policy_basis"] = []
            row["meal_policy_error"] = error or "meal policy could not be classified"
            continue
        row["meal_cap_policy"] = policy["policy"]
        row["meal_daily_cap"] = money(policy["cap"])
        row["meal_policy_basis"] = policy["basis"]
        row["meal_policy_error"] = ""


def meal_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "user_no": row.get("user_no", ""),
        "proof_no": row.get("proof_no", ""),
        "source_unit_id": row.get("source_unit_id", ""),
        "source_filename": row.get("source_filename") or row.get("supporting_invoice_filename") or "",
        "seller_name": row.get("seller_name", ""),
        "invoice_amount": row.get("invoice_amount", "0.00"),
        "reimbursable_amount": row.get("reimbursable_amount", row.get("amount", "0.00")),
        "attendees": row.get("attendees", ""),
        "note": row.get("note", ""),
        "amount_column": row.get("amount_column", ""),
        "expenses_nature": row.get("expenses_nature", ""),
        "meal_cap_policy": row.get("meal_cap_policy", ""),
        "meal_daily_cap": row.get("meal_daily_cap", ""),
        "meal_policy_basis": row.get("meal_policy_basis", []),
    }


def suggest_meal_adjustments(day_rows: list[dict[str, Any]], over_by: Decimal) -> list[dict[str, Any]]:
    remaining = over_by
    adjustments: list[dict[str, Any]] = []
    candidates = sorted(
        day_rows,
        key=lambda row: (Decimal(money(row.get("amount"))), clean(row.get("source_unit_id"))),
        reverse=True,
    )
    for row in candidates:
        if remaining <= 0:
            break
        current = Decimal(money(row.get("amount")))
        if current <= 0:
            continue
        reduction = min(current, remaining)
        suggested = current - reduction
        adjustments.append({
            "user_no": row.get("user_no", ""),
            "proof_no": row.get("proof_no", ""),
            "source_unit_id": row.get("source_unit_id", ""),
            "source_filename": row.get("source_filename") or row.get("supporting_invoice_filename") or "",
            "current_reimbursable_amount": money(current),
            "suggested_reimbursable_amount": money(suggested),
            "reduce_by": money(reduction),
        })
        remaining -= reduction
    return adjustments


def meal_daily_cap_checks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_policy_date: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    policies: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        policy = meal_cap_policy(row)
        if not policy:
            continue
        date_value = row.get("date") or date_yyyymmdd(row.get("expense_date", ""))
        key = (policy["policy"], date_value)
        by_policy_date[key].append(row)
        policies[key] = policy

    checks: list[dict[str, Any]] = []
    for key, day_rows in sorted(by_policy_date.items(), key=lambda item: (item[0][1], item[0][0])):
        policy = policies[key]
        date_value = key[1]
        cap = policy["cap"]
        total = sum((Decimal(money(row.get("amount"))) for row in day_rows), Decimal("0.00"))
        over_by = total - cap
        has_attendees = any(clean(row.get("attendees")) for row in day_rows)
        status = C["meal_cap_ok"]
        severity = "ok"
        requires_confirmation = False
        suggestions: list[dict[str, Any]] = []
        if over_by > 0:
            if has_attendees:
                status = C["meal_cap_over_with_attendees"]
                severity = "advisory"
            else:
                status = C["meal_cap_over_needs_confirmation"]
                severity = "blocking"
                requires_confirmation = True
                suggestions = suggest_meal_adjustments(day_rows, over_by)
        event_declared = str(policy["policy"]).startswith("event_declared")
        event_direction = next(
            (clean(row.get("event_meal_direction")) for row in day_rows if clean(row.get("event_meal_direction"))),
            "",
        )
        checks.append({
            "policy": policy["policy"],
            "policy_name": policy["policy_name"],
            "date": date_value,
            "cap": money(cap),
            "event_declared": event_declared,
            "event_meal_direction": event_direction,
            "aggregation_key": "meal_cap_policy + expense_date",
            "cross_column_aggregation": True,
            "policy_basis": sorted({
                basis
                for row in day_rows
                for basis in row.get("meal_policy_basis", [])
            }),
            "policy_non_triggers": list(MEAL_POLICY_NON_TRIGGERS),
            "total": money(total),
            "over_by": money(over_by) if over_by > 0 else "0.00",
            "status": status,
            "severity": severity,
            "advisory": severity == "advisory",
            "has_attendees": has_attendees,
            "requires_user_confirmation": requires_confirmation,
            "items": [meal_item(row) for row in day_rows],
            "suggested_adjustments": suggestions,
        })
    return checks


def _meal_population_line(meal_checks: list) -> str:
    total = sum(len(day.get("items", [])) for day in meal_checks)
    return (
        f"(covering {total} meal item(s) across {len(meal_checks)} policy/day pool(s); "
        "source_category=meal selects the population only)"
    )


def meal_policy_rule() -> dict[str, Any]:
    return {
        "population_selector": "source_category=meal",
        "cap_selector": "final_note/meal_context substantive purpose",
        "business_trip_meal": {
            "cap": money(BUSINESS_TRIP_MEAL_DAILY_CAP),
            "signals": ["final_note starts with 出差餐费", "meal_context=business_trip/station_airport/travel"],
        },
        "local_overtime_meal": {
            "cap": money(LOCAL_OVERTIME_MEAL_DAILY_CAP),
            "signals": ["final_note starts with 加班餐费", "meal_context=overtime"],
        },
        "not_cap_selectors": list(MEAL_POLICY_NON_TRIGGERS),
        "cross_column_aggregation": (
            "Same-date rows with the same meal_cap_policy are summed together even when one is in meal "
            "and another is in travel."
        ),
        "critical_invariant": (
            "A Shanghai row may use amount_column=meal and expenses_nature=本地 while remaining "
            "business_trip_meal with a 150/day cap. There is no generic Shanghai/local meal 60/day rule; "
            "60/day applies only to explicit local overtime meals."
        ),
    }


def print_meal_cap_check(checks: list[dict[str, Any]]) -> None:
    print("\n=== MEAL DAILY CAP CHECK TO RELAY VERBATIM ===")
    print(
        "POLICY INVARIANT: source_category=meal only selects meal rows. The 150/60 cap is selected only "
        "by final_note/meal_context: 出差餐费=150/day; explicit 加班餐费=60/day."
    )
    print(
        "NEVER select a cap from city, amount_column, or Expense Nature. A Shanghai row in the meal column "
        "with Note=出差餐费 is still in the 150/day business-trip pool. There is no generic 本地餐=60 rule."
    )
    print(
        "AGGREGATION: sum all rows sharing meal_cap_policy + date, including rows split across the meal and "
        "travel columns. Do not recalculate separate pools by workbook column."
    )
    print(_meal_population_line(checks))
    print("Relay this block as generated; do not independently reclassify or recompute it from workbook columns.")
    if not checks:
        print("No meal rows requiring daily cap checks found.")
        print("=== END MEAL DAILY CAP CHECK ===")
        return
    for check in checks:
        print(
            f"- {check['date']} [{check.get('policy_name', '')}]: total {check['total']} / cap {check['cap']} / "
            f"over {check['over_by']} / {check['status']} / severity {check.get('severity', 'ok')}"
        )
        for item in check["items"]:
            print(
                f"  item {item.get('user_no') or '-'} | proof {item.get('proof_no') or '-'} | "
                f"{item.get('source_filename') or '-'} | invoice {item['invoice_amount']} | "
                f"reimburse {item['reimbursable_amount']} | policy {item.get('meal_cap_policy') or '-'} "
                f"cap {item.get('meal_daily_cap') or '-'} | column {item.get('amount_column') or '-'} | "
                f"nature {item.get('expenses_nature') or '-'} | note {item.get('note') or '-'} | "
                f"attendees {item.get('attendees') or '-'}"
            )
        if check["suggested_adjustments"]:
            print("  suggested adjustment to fit cap:")
            for adjustment in check["suggested_adjustments"]:
                print(
                    f"  item {adjustment.get('user_no') or '-'} "
                    f"{adjustment['current_reimbursable_amount']} -> "
                    f"{adjustment['suggested_reimbursable_amount']} "
                    f"(reduce {adjustment['reduce_by']})"
                )
    print("=== END MEAL DAILY CAP CHECK ===")


def boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return clean(value).lower() in {"1", "true", "yes", "y", "provided", "\u662f", "\u6709", "\u540c\u4f4f"}


def parse_positive_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    match = re.search(r"\d+", str(value))
    if not match:
        return None
    parsed = int(match.group(0))
    return parsed if parsed > 0 else None


def parse_date_object(value: Any) -> date | None:
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", clean(value))
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
    except ValueError:
        return None


def date_strings(value: Any) -> list[str]:
    return [
        f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
        for match in re.finditer(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", clean(value))
    ]


def hotel_city_policy(row: dict[str, Any]) -> dict[str, Any] | None:
    tier = clean(row.get("hotel_city_tier")).lower()
    if tier in {"first_tier", "tier1", "1", "\u4e00\u7ebf", "\u5317\u4e0a\u5e7f\u6df1"}:
        return {"city_tier": "first_tier", "policy_name": C["hotel_policy_first_tier"], "cap_per_night": FIRST_TIER_HOTEL_CAP}
    if tier in {"other", "non_first_tier", "2", "\u5176\u4ed6", "\u975e\u4e00\u7ebf"}:
        return {"city_tier": "other", "policy_name": C["hotel_policy_other_city"], "cap_per_night": OTHER_CITY_HOTEL_CAP}
    city = clean(row.get("hotel_city") or row.get("city"))
    if not city:
        return None
    city_lower = city.lower()
    if any(name in city or name in city_lower for name in FIRST_TIER_CITIES):
        return {"city_tier": "first_tier", "policy_name": C["hotel_policy_first_tier"], "cap_per_night": FIRST_TIER_HOTEL_CAP}
    return {"city_tier": "other", "policy_name": C["hotel_policy_other_city"], "cap_per_night": OTHER_CITY_HOTEL_CAP}


def hotel_stay_details(row: dict[str, Any]) -> dict[str, Any]:
    note = clean(row.get("note"))
    nights = parse_positive_int(row.get("hotel_nights") or row.get("nights"))
    if not nights:
        night_match = re.search(r"(\d+)\s*\u665a", note)
        if night_match:
            nights = int(night_match.group(1))
    check_in = clean(row.get("check_in_date"))
    check_out = clean(row.get("check_out_date"))
    date_candidates = date_strings(note)
    if not check_in and date_candidates:
        check_in = date_candidates[0]
    if not check_out and len(date_candidates) >= 2:
        check_out = date_candidates[1]
    if not nights and check_in and check_out:
        start = parse_date_object(check_in)
        end = parse_date_object(check_out)
        if start and end and end > start:
            nights = (end - start).days
    return {"nights": nights, "check_in_date": check_in, "check_out_date": check_out}


def has_hotel_shared_room(row: dict[str, Any]) -> bool:
    return boolish(row.get("shared_room")) or bool(clean(row.get("room_shared_with") or row.get("room_share_note") or row.get("attendees")))


def hotel_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "user_no": row.get("user_no", ""),
        "proof_no": row.get("proof_no", ""),
        "source_unit_id": row.get("source_unit_id", ""),
        "source_filename": row.get("source_filename") or row.get("supporting_invoice_filename") or "",
        "seller_name": row.get("seller_name", ""),
        "invoice_amount": row.get("invoice_amount", "0.00"),
        "reimbursable_amount": row.get("reimbursable_amount", row.get("amount", "0.00")),
        "note": row.get("note", ""),
        "room_shared_with": row.get("room_shared_with", ""),
        "room_share_note": row.get("room_share_note", ""),
    }


def hotel_cap_checks(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for row in rows:
        if row.get("source_category") != "hotel" and row.get("amount_column") != "hotel":
            continue
        stay = hotel_stay_details(row)
        policy = hotel_city_policy(row)
        amount = Decimal(money(row.get("amount")))
        has_shared = has_hotel_shared_room(row)
        status = C["meal_cap_ok"]
        severity = "ok"
        requires_confirmation = False
        suggested_adjustments: list[dict[str, Any]] = []
        cap_total: Decimal | None = None
        over_by: Decimal | None = None

        if not stay["nights"]:
            status = C["hotel_cap_missing_nights"]
            severity = "blocking"
            requires_confirmation = True
        elif not policy:
            status = C["hotel_cap_missing_city"]
            severity = "blocking"
            requires_confirmation = True
        else:
            cap_total = policy["cap_per_night"] * Decimal(stay["nights"])
            over_by = amount - cap_total
            if over_by > 0:
                if has_shared:
                    status = C["hotel_cap_over_with_shared_room"]
                    severity = "advisory"
                else:
                    status = C["hotel_cap_over_needs_confirmation"]
                    severity = "blocking"
                    requires_confirmation = True
                    suggested_adjustments.append({
                        "user_no": row.get("user_no", ""),
                        "proof_no": row.get("proof_no", ""),
                        "source_unit_id": row.get("source_unit_id", ""),
                        "source_filename": row.get("source_filename") or row.get("supporting_invoice_filename") or "",
                        "current_reimbursable_amount": money(amount),
                        "suggested_reimbursable_amount": money(cap_total),
                        "reduce_by": money(over_by),
                    })

        checks.append({
            "policy": "hotel_cap",
            "policy_name": policy["policy_name"] if policy else "",
            "city_tier": policy["city_tier"] if policy else "",
            "city": clean(row.get("hotel_city") or row.get("city")),
            "date": row.get("date") or date_yyyymmdd(row.get("expense_date", "")),
            "check_in_date": stay["check_in_date"],
            "check_out_date": stay["check_out_date"],
            "nights": stay["nights"] or "",
            "cap_per_night": money(policy["cap_per_night"]) if policy else "",
            "cap_total": money(cap_total) if cap_total is not None else "",
            "total": money(amount),
            "over_by": money(over_by) if over_by and over_by > 0 else "0.00",
            "status": status,
            "severity": severity,
            "advisory": severity == "advisory",
            "has_shared_room": has_shared,
            "requires_user_confirmation": requires_confirmation,
            "items": [hotel_item(row)],
            "suggested_adjustments": suggested_adjustments,
        })
    return checks


def print_hotel_cap_check(checks: list[dict[str, Any]]) -> None:
    print("\nHOTEL CAP CHECK TO SHOW IN CHAT")
    print("Copy or summarize this check in the conversation before final submission.")
    if not checks:
        print("No hotel rows requiring cap checks found.")
        return
    for check in checks:
        print(
            f"- item {check['items'][0].get('user_no') or '-'} | {check.get('city') or '-'} "
            f"[{check.get('policy_name') or 'city tier unknown'}] | "
            f"{check.get('nights') or '?'} night(s) | total {check['total']} / "
            f"cap {check.get('cap_total') or '?'} / over {check['over_by']} / {check['status']} / "
            f"severity {check.get('severity', 'ok')}"
        )
        item = check["items"][0]
        print(
            f"  proof {item.get('proof_no') or '-'} | {item.get('source_filename') or '-'} | "
            f"invoice {item['invoice_amount']} | reimburse {item['reimbursable_amount']} | "
            f"shared {item.get('room_shared_with') or item.get('room_share_note') or '-'}"
        )
        if check["suggested_adjustments"]:
            print("  suggested adjustment to fit cap:")
            for adjustment in check["suggested_adjustments"]:
                print(
                    f"  item {adjustment.get('user_no') or '-'} "
                    f"{adjustment['current_reimbursable_amount']} -> "
                    f"{adjustment['suggested_reimbursable_amount']} "
                    f"(reduce {adjustment['reduce_by']})"
                )


def copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    for col in range(1, 14):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)


def font(name: str, bold: bool = False, size: int = 9, color: str | None = None) -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def fill(color: str | None = None) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color) if color else PatternFill(fill_type=None)


def column_count(layout: dict[str, Any]) -> int:
    return int(layout["columns"]["count"])


def style_columns(layout: dict[str, Any], key: str) -> set[int]:
    return {int(value) for value in layout["styles"].get(key, [])}


def sample_rows(layout: dict[str, Any] | None) -> dict[str, int]:
    if layout:
        return {name: int(row) for name, row in layout["sample_rows"].items()}
    return {"detail": 3, "subtotal": 6, "column_summary": 20, "total": 21, "grand": 22, "status": 23}


def column_format(layout: dict[str, Any], col_idx: int, *, text_default: bool = False) -> str:
    formats = layout["formats"]
    if col_idx in style_columns(layout, "amount_columns"):
        return formats["money"]
    if col_idx == int(layout["styles"]["proof_no_column"]):
        return formats["integer"]
    if text_default or col_idx in style_columns(layout, "text_columns"):
        return formats["text"]
    return formats["general"]


def border_for(
    col_idx: int,
    *,
    column_total: int = 13,
    top: str | None = "thin",
    bottom: str | None = "thin",
    inner: str | None = "thin",
) -> Border:
    return Border(
        left=Side(style="medium" if col_idx == 1 else None),
        right=Side(style="medium" if col_idx == column_total else inner),
        top=Side(style=top),
        bottom=Side(style=bottom),
    )


def set_cell_style(
    cell: Any,
    *,
    layout: dict[str, Any],
    font_name: str | None = None,
    bold: bool = False,
    font_color: str | None = None,
    fill_color: str | None = None,
    horizontal: str = "center",
    vertical: str = "center",
    wrap_text: bool = False,
    number_format: str = "General",
    border: Border | None = None,
) -> None:
    fonts = layout["fonts"]
    cell.font = font(font_name or fonts["latin"], bold=bold, size=int(fonts["size"]), color=font_color)
    cell.fill = fill(fill_color)
    cell.alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    cell.number_format = number_format
    cell.border = border or border_for(cell.column, column_total=column_count(layout))


def style_instruction_row(ws: Any, layout: dict[str, Any]) -> None:
    ws.row_dimensions[1].height = float(layout["rows"]["instruction_height"])
    fonts = layout["fonts"]
    colors = layout["colors"]
    latin_columns = style_columns(layout, "latin_instruction_columns")
    for col_idx, value in enumerate(layout["instruction_row"]["values"], start=1):
        cell = ws.cell(1, col_idx)
        cell.value = value
        set_cell_style(
            cell,
            layout=layout,
            font_name=fonts["latin"] if col_idx in latin_columns else fonts["cjk"],
            fill_color=colors["instruction_fill"],
            wrap_text=True,
            number_format=column_format(layout, col_idx),
            border=border_for(col_idx, column_total=column_count(layout), top="medium", bottom="thin"),
        )


def style_header_row(ws: Any, layout: dict[str, Any]) -> None:
    ws.row_dimensions[2].height = float(layout["rows"]["header_height"])
    colors = layout["colors"]
    date_header_column = int(layout["styles"]["date_header_column"])
    for col_idx, value in enumerate(layout["header_row"]["values"], start=1):
        cell = ws.cell(2, col_idx)
        cell.value = value
        set_cell_style(
            cell,
            layout=layout,
            bold=True,
            font_color=colors["date_header_font"] if col_idx == date_header_column else None,
            wrap_text=True,
            number_format=column_format(layout, col_idx, text_default=True),
            border=Border(
                left=Side(style="medium" if col_idx == 1 else None),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium"),
            ),
        )


def style_detail_row(ws: Any, row: int, layout: dict[str, Any]) -> None:
    ws.row_dimensions[row].height = float(layout["rows"]["detail_height"])
    fonts = layout["fonts"]
    cjk_columns = style_columns(layout, "cjk_detail_columns")
    left_columns = style_columns(layout, "left_aligned_detail_columns")
    wrapped_columns = style_columns(layout, "wrapped_detail_columns")
    proof_no_column = int(layout["styles"]["proof_no_column"])
    for col_idx in range(1, column_count(layout) + 1):
        set_cell_style(
            ws.cell(row, col_idx),
            layout=layout,
            font_name=fonts["cjk"] if col_idx in cjk_columns else fonts["latin"],
            bold=col_idx == proof_no_column,
            horizontal="left" if col_idx in left_columns else "center",
            wrap_text=col_idx in wrapped_columns,
            number_format=column_format(layout, col_idx),
        )


def style_subtotal_row(ws: Any, row: int, layout: dict[str, Any]) -> None:
    ws.row_dimensions[row].height = float(layout["rows"]["subtotal_height"])
    fonts = layout["fonts"]
    colors = layout["colors"]
    label_column = int(layout["styles"]["subtotal_label_column"])
    formula_column = int(layout["styles"]["subtotal_formula_column"])
    bold_columns = style_columns(layout, "subtotal_bold_columns")
    for col_idx in range(1, column_count(layout) + 1):
        number_format = layout["formats"]["money"] if col_idx == formula_column else layout["formats"]["general"]
        set_cell_style(
            ws.cell(row, col_idx),
            layout=layout,
            font_name=fonts["cjk"] if col_idx == label_column else fonts["latin"],
            bold=col_idx in bold_columns,
            fill_color=colors["subtotal_fill"],
            number_format=number_format,
            border=border_for(col_idx, column_total=column_count(layout), top=None, bottom="thin"),
        )


def style_summary_row(ws: Any, row: int, kind: str, layout: dict[str, Any]) -> None:
    ws.row_dimensions[row].height = float(layout["rows"]["summary_height"])
    formula_column = int(layout["styles"]["summary_formula_column"])
    amount_columns = style_columns(layout, "amount_columns")
    for col_idx in range(1, column_count(layout) + 1):
        number_format = layout["formats"]["general"]
        if (kind == "column_summary" and col_idx in amount_columns) or (kind in {"total", "grand"} and col_idx == formula_column):
            number_format = layout["formats"]["money"]
        set_cell_style(
            ws.cell(row, col_idx),
            layout=layout,
            bold=col_idx >= 5,
            number_format=number_format,
            border=Border(),
        )


def create_base_workbook(layout: dict[str, Any]) -> Any:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = layout["sheet"]["name"]
    ws.sheet_view.showGridLines = bool(layout["sheet"].get("show_grid_lines", True))
    for col_letter, width in layout["columns"]["widths"].items():
        ws.column_dimensions[col_letter].width = width
    rows = sample_rows(layout)
    style_instruction_row(ws, layout)
    style_header_row(ws, layout)
    style_detail_row(ws, rows["detail"], layout)
    style_subtotal_row(ws, rows["subtotal"], layout)
    style_summary_row(ws, rows["column_summary"], "column_summary", layout)
    style_summary_row(ws, rows["total"], "total", layout)
    style_summary_row(ws, rows["grand"], "grand", layout)
    style_summary_row(ws, rows["status"], "status", layout)
    labels = layout["labels"]
    ws.cell(rows["subtotal"], int(layout["styles"]["subtotal_label_column"])).value = labels["project_subtotal"]
    ws.cell(rows["subtotal"], int(layout["styles"]["subtotal_formula_column"])).value = 0
    ws.cell(rows["total"], int(layout["styles"]["summary_label_column"])).value = labels["total"]
    ws.cell(rows["total"], int(layout["styles"]["summary_formula_column"])).value = 0
    ws.cell(rows["grand"], int(layout["styles"]["summary_label_column"])).value = labels["grand_total"]
    ws.cell(rows["grand"], int(layout["styles"]["summary_formula_column"])).value = 0
    ws.cell(rows["status"], int(layout["styles"]["summary_label_column"])).value = labels["status"]
    ws.cell(rows["status"], int(layout["styles"]["summary_formula_column"])).value = labels["initial_status"]
    return wb


def capture_styles(ws: Any, layout: dict[str, Any] | None = None) -> dict[str, list[Any]]:
    styles = {}
    total_columns = column_count(layout) if layout else 13
    for name, row in sample_rows(layout).items():
        styles[name] = []
        for col in range(1, total_columns + 1):
            cell = ws.cell(row, col)
            styles[name].append({
                "style": copy.copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy.copy(cell.alignment),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "font": copy.copy(cell.font),
            })
    return styles


def apply_style(ws: Any, row: int, style_row: list[Any]) -> None:
    for col, style in enumerate(style_row, start=1):
        cell = ws.cell(row, col)
        cell._style = copy.copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy.copy(style["alignment"])
        cell.border = copy.copy(style["border"])
        cell.fill = copy.copy(style["fill"])
        cell.font = copy.copy(style["font"])


def write_workbook(
    output: Path,
    rows: list[dict[str, Any]],
    template: Path | None = None,
    layout: dict[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    if not template and layout is None:
        layout = load_layout(bundled_layout_path())
    wb = openpyxl.load_workbook(template) if template else create_base_workbook(layout)
    ws = wb.active
    styles = capture_styles(ws, layout)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    project_groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for row in rows:
        key = f"{row['client']}|{row['client_charge_code']}"
        project_groups.setdefault(key, []).append(row)

    output_row = 3
    project_blocks = []
    subtotal_rows = []
    for key, group_rows in project_groups.items():
        group_rows.sort(key=lambda r: (ROW_ORDER.get(r["row_order_type"], 99), r["expense_date"], r["proof_no"]))
        first = output_row
        for row_data in group_rows:
            apply_style(ws, output_row, styles["detail"])
            values = [
                row_data["date"],
                row_data["requester"],
                row_data["client"],
                row_data["client_charge_code"],
                row_data["expenses_nature"],
                row_data["note"],
            ]
            for col_idx, value in enumerate(values, start=1):
                ws.cell(output_row, col_idx).value = value
            for col_name, col_letter in AMOUNT_COLUMNS.items():
                ws[f"{col_letter}{output_row}"] = Decimal(row_data["amount"]) if row_data["amount_column"] == col_name else None
            ws.cell(output_row, 13).value = row_data["proof_no"]
            row_data["excel_row"] = output_row
            output_row += 1
        last = output_row - 1
        subtotal = output_row
        apply_style(ws, subtotal, styles["subtotal"])
        subtotal_label = layout["labels"]["project_subtotal"] if layout else C["summary"]
        ws[f"D{subtotal}"] = subtotal_label
        ws[f"F{subtotal}"] = f"=SUM(G{first}:L{last})"
        subtotal_rows.append(subtotal)
        client, code = key.split("|", 1)
        project_blocks.append({
            "project_key": key,
            "client": client,
            "client_charge_code": code,
            "first_detail_row": first,
            "last_detail_row": last,
            "subtotal_row": subtotal,
            "subtotal_formula": f"SUM(G{first}:L{last})",
        })
        output_row += 1

    last_project_subtotal_row = output_row - 1
    column_summary_row = output_row
    apply_style(ws, column_summary_row, styles["column_summary"])
    for col_letter in AMOUNT_COLUMNS.values():
        ws[f"{col_letter}{column_summary_row}"] = f"=SUM({col_letter}3:{col_letter}{last_project_subtotal_row})"

    total_row = output_row + 1
    apply_style(ws, total_row, styles["total"])
    ws[f"E{total_row}"] = layout["labels"]["total"] if layout else "Total: (RMB)"
    ws[f"F{total_row}"] = f"=SUM(G{column_summary_row}:L{column_summary_row})"

    grand_total_row = output_row + 2
    apply_style(ws, grand_total_row, styles["grand"])
    ws[f"E{grand_total_row}"] = layout["labels"]["grand_total"] if layout else "Grand Total: (RMB)"
    refs = ",".join(f"F{r}" for r in subtotal_rows) or "0"
    ws[f"F{grand_total_row}"] = f"=SUM({refs})"

    status_row = output_row + 3
    apply_style(ws, status_row, styles["status"])
    ws[f"E{status_row}"] = layout["labels"]["status"] if layout else "Status"
    ws[f"F{status_row}"] = f"=F{total_row}=F{grand_total_row}"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    summary_rows = {
        "column_summary_row": column_summary_row,
        "total_row": total_row,
        "grand_total_row": grand_total_row,
        "status_row": status_row,
    }
    return project_blocks, summary_rows


def build_markdown(payload: dict[str, Any], workbook: Path) -> str:
    lines = [
        "# Final Expense Rows",
        "",
        f"Generated at: {payload['generated_at']}",
        f"Requester: {payload['requester']}",
        f"Workbook: {workbook}",
        "",
        "## Rows",
        "",
        "| Excel Row | Date | Client | Code | Nature | Note | Column | Amount | No. |",
        "| ---: | --- | --- | --- | --- | --- | --- | ---: | ---: |",
    ]
    for row in payload["rows"]:
        lines.append(
            f"| {row.get('excel_row','')} | {row['date']} | {row['client']} | {row['client_charge_code']} | "
            f"{row['expenses_nature']} | {row['note']} | {row['amount_column']} | {row['amount']} | {row['proof_no']} |"
        )
    lines += ["", "## Project Blocks", "", "| Project | Rows | Subtotal Row | Formula |", "| --- | --- | ---: | --- |"]
    for block in payload["project_blocks"]:
        lines.append(
            f"| {block['project_key']} | {block['first_detail_row']}:{block['last_detail_row']} | "
            f"{block['subtotal_row']} | {block['subtotal_formula']} |"
        )
    chains = payload.get("rail_journey_chains", [])
    if chains:
        lines += [
            "",
            "## Railway Journey Chains",
            "",
            "| Chain | Items | Route | Project | Code | Rule | Needs Confirmation |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
        for chain in chains:
            lines.append(
                f"| {chain.get('journey_chain_id', '')} | {', '.join(str(value) for value in chain.get('user_nos', []))} | "
                f"{chain.get('route', '')} | {chain.get('client_name', '')} | {chain.get('client_charge_code', '')} | "
                f"{chain.get('assignment_rule', '')} | {chain.get('needs_confirmation', False)} |"
            )
    hint_records = payload.get("expense_hint_reconciliation", [])
    if hint_records:
        lines += [
            "",
            "## User Expense Record Reconciliation",
            "",
            "| Hint | Record | Match | Resolution | Matched Items | Answer |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
        for record in hint_records:
            lines.append(
                f"| {record.get('hint_id', '')} | {record.get('summary', '')} | "
                f"{record.get('match_status', '')} | {record.get('resolution_status', '')} | "
                f"{', '.join(str(value) for value in record.get('matched_user_nos', []))} | "
                f"{record.get('resolution_answer', '')} |"
            )
    meal_rule = payload.get("meal_policy_rule", {})
    lines += [
        "",
        "## Meal Policy Rule",
        "",
        f"- Population selector: {meal_rule.get('population_selector', '')}",
        f"- Cap selector: {meal_rule.get('cap_selector', '')}",
        f"- Not cap selectors: {', '.join(meal_rule.get('not_cap_selectors', []))}",
        f"- Invariant: {meal_rule.get('critical_invariant', '')}",
        f"- Aggregation: {meal_rule.get('cross_column_aggregation', '')}",
    ]
    lines += [
        "",
        "## Meal Daily Cap Checks",
        "",
        "| Policy | Date | Total | Cap | Over By | Status | Severity | Needs Confirmation |",
        "| --- | --- | ---: | ---: | ---: | --- | --- | --- |",
    ]
    for check in payload.get("meal_daily_cap_checks", []):
        lines.append(
            f"| {check.get('policy_name', '')} | {check['date']} | {check['total']} | {check['cap']} | {check['over_by']} | "
            f"{check['status']} | {check.get('severity', 'ok')} | {check['requires_user_confirmation']} |"
        )
    lines += [
        "",
        "## Hotel Cap Checks",
        "",
        "| Item | City | Nights | Total | Cap Total | Over By | Status | Severity | Needs Confirmation |",
        "| ---: | --- | ---: | ---: | ---: | ---: | --- | --- | --- |",
    ]
    for check in payload.get("hotel_cap_checks", []):
        item = check.get("items", [{}])[0]
        lines.append(
            f"| {item.get('user_no', '')} | {check.get('city', '')} | {check.get('nights', '')} | "
            f"{check['total']} | {check.get('cap_total', '')} | {check['over_by']} | "
            f"{check['status']} | {check.get('severity', 'ok')} | {check['requires_user_confirmation']} |"
        )
    return "\n".join(lines) + "\n"


def aggregate_check_status(checks: list[dict[str, Any]]) -> str:
    if any(check.get("requires_user_confirmation") for check in checks):
        return "needs_confirmation"
    if any(check.get("severity") == "advisory" or check.get("advisory") for check in checks):
        return "advisory"
    return "ok"


def checks_requiring_confirmation(payload: dict[str, Any]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    checks.extend(check for check in payload.get("meal_daily_cap_checks", []) if check.get("requires_user_confirmation"))
    checks.extend(check for check in payload.get("hotel_cap_checks", []) if check.get("requires_user_confirmation"))
    return checks


def advisory_checks(payload: dict[str, Any]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    checks.extend(
        check for check in payload.get("meal_daily_cap_checks", [])
        if check.get("severity") == "advisory" or check.get("advisory")
    )
    checks.extend(
        check for check in payload.get("hotel_cap_checks", [])
        if check.get("severity") == "advisory" or check.get("advisory")
    )
    return checks


def print_stage3_review_summary(payload: dict[str, Any]) -> None:
    blocking = checks_requiring_confirmation(payload)
    advisory = advisory_checks(payload)
    print("\nSTAGE 3 REVIEW SUMMARY TO SHOW IN CHAT")
    print(
        "AUTHORITATIVE MEAL RESULT: use the generated MEAL DAILY CAP CHECK block and per-row "
        "meal_cap_policy/meal_daily_cap fields; never reclassify from city, amount column, or Expense Nature."
    )
    print(
        "meal_daily_caps="
        f"{payload['checks'][0]['status']} "
        f"({payload['checks'][0]['days_requiring_confirmation']} requiring confirmation, "
        f"{payload['checks'][0]['days_with_advisory']} advisory)"
    )
    print(
        "hotel_caps="
        f"{payload['checks'][1]['status']} "
        f"({payload['checks'][1]['items_requiring_confirmation']} requiring confirmation, "
        f"{payload['checks'][1]['items_with_advisory']} advisory)"
    )
    if blocking:
        print("ACTION REQUIRED: The workbook was written, but final submission is blocked until these checks are confirmed in chat.")
        for check in blocking:
            label = check.get("policy_name") or check.get("policy") or "policy_check"
            date_or_item = check.get("date") or (check.get("items") or [{}])[0].get("user_no") or "-"
            print(
                f"- {label} {date_or_item}: total {check.get('total')} / "
                f"cap {check.get('cap') or check.get('cap_total') or '?'} / "
                f"over {check.get('over_by')} / {check.get('status')}"
            )
        return
    if advisory:
        print("ADVISORY: No blocking cap issue, but the following checks should still be summarized for the applicant.")
        for check in advisory:
            label = check.get("policy_name") or check.get("policy") or "policy_check"
            date_or_item = check.get("date") or (check.get("items") or [{}])[0].get("user_no") or "-"
            print(
                f"- {label} {date_or_item}: total {check.get('total')} / "
                f"cap {check.get('cap') or check.get('cap_total') or '?'} / "
                f"over {check.get('over_by')} / {check.get('status')}"
            )
        return
    print(
        "OK: No meal or hotel cap issue requiring applicant attention. This means zero blocking/advisory "
        "issues, not zero checked meal rows; retain the generated policy classification and totals."
    )


def main(argv: list[str] | None = None) -> int:
    configure_stdio()
    parser = argparse.ArgumentParser(description="Write reimbursement workbook from stage-2 allocation JSON.")
    parser.add_argument("--allocation", required=True, help="Path to process/expense-allocation.json.")
    parser.add_argument(
        "--template",
        help="Optional reimbursement template .xlsx, or 'bundled' for assets/reimbursement-template.xlsx. If omitted, the workbook is generated directly by script.",
    )
    parser.add_argument(
        "--layout",
        help="Optional generated-workbook layout TOML. Defaults to assets/reimbursement-workbook-layout.toml when no template is supplied.",
    )
    parser.add_argument("--output", required=True, help="Output .xlsx path.")
    parser.add_argument("--requester", required=True, help="Requester name.")
    parser.add_argument("--process-dir", default="process", help="Folder for final-expense-rows outputs.")
    parser.add_argument("--allow-unconfirmed", action="store_true", help="Allow writing with open questions/unconfirmed units.")
    args = parser.parse_args(argv)

    allocation_path = Path(args.allocation)
    process_dir = Path(args.process_dir)
    if process_dir.resolve() != allocation_path.parent.resolve():
        print(
            "ERROR: --process-dir must be the canonical directory containing --allocation. "
            "A different process folder could hide the current independent-review gate or mix batches. "
            f"Expected: {allocation_path.parent.resolve()}",
            file=sys.stderr,
        )
        return ExitCode.COMMAND_ERROR
    template_path = resolve_template_arg(args.template)
    if template_path and not template_path.exists():
        print(
            f"ERROR: Template workbook not found: {template_path}. "
            "Pass an existing --template path or omit --template to generate the workbook directly.",
            file=sys.stderr,
        )
        return ExitCode.COMMAND_ERROR
    layout_path = resolve_layout_arg(args.layout)
    layout = None
    if not template_path or args.layout:
        try:
            layout = load_layout(layout_path)
        except (OSError, ValueError, tomllib.TOMLDecodeError) as exc:
            print(f"ERROR: Workbook layout could not be loaded from {layout_path}: {exc}", file=sys.stderr)
            return ExitCode.COMMAND_ERROR

    allocation = load_json(allocation_path)
    integrity.require_valid(allocation, allocation_path)
    for unit in allocation.get("allocation_units", []):
        if unit.get("is_substitute_invoice") and unit.get("approval_file"):
            evidence_paths.normalize_approval_file(unit, process_dir)
    errors = require_ready(allocation, args.allow_unconfirmed)
    expected_context_sha = str(allocation.get("source_project_context_sha256", "")).strip()
    recorded_context = str(allocation.get("source_project_context_file", "")).strip()
    if expected_context_sha:
        context_path = Path(recorded_context).expanduser() if recorded_context else Path()
        if recorded_context and not context_path.is_absolute():
            context_path = allocation_path.parent.parent / context_path
        if not recorded_context or not context_path.is_file():
            errors.append(
                "The project context used by allocation is missing. Restore/rewrite the canonical "
                "project-context.json and rerun Stage 2 plus Composer before writing Excel."
            )
        else:
            try:
                actual_context_sha = hashlib.sha256(context_path.read_bytes()).hexdigest()
            except OSError as exc:
                errors.append(f"The project context cannot be read for provenance validation: {exc}")
            else:
                if actual_context_sha != expected_context_sha:
                    errors.append(
                        "Project context changed after allocation was created. Rerun Stage 2, recompose "
                        "answers, and apply them before writing Excel."
                    )
    allocation_text_issues = text_safety.find_suspect_text(
        {
            "allocation_units": text_safety.pick_fields(
                allocation.get("allocation_units", []), ALLOCATION_TEXT_FIELDS
            )
        },
        path="allocation",
    )
    if allocation_text_issues:
        errors.append(
            "Allocation contains suspect encoding damage in user-visible fields. Correct the UTF-8 "
            "allocation_decisions.v1 input, rerun Composer, and apply it through the updater: "
            + "; ".join(allocation_text_issues)
        )
    extraction_path = allocation_path.parent / "invoice-extraction.json"
    if not extraction_path.exists():
        errors.append(
            f"Current extraction not found next to allocation: {extraction_path}. "
            "Re-run Stage 1 and Stage 2 from the same process folder before writing Excel."
        )
    else:
        extraction = load_json(extraction_path)
        integrity.require_valid(extraction, extraction_path, kind="extraction")
        extraction_fp = (extraction.get("integrity") or {}).get("fingerprint", "")
        allocated_fp = str(allocation.get("source_extraction_fingerprint", ""))
        if not allocated_fp:
            errors.append(
                "Allocation has no source_extraction_fingerprint and cannot prove which extraction generation "
                "it used. Re-run allocate_expenses.py and recompose decisions."
            )
        elif allocated_fp != extraction_fp:
            errors.append(
                "Extraction changed after allocation was created. Re-run allocate_expenses.py, recompose "
                "decisions, and reapply the user's answers before writing Excel."
            )
        unresolved_inputs = [
            item for item in extraction.get("unresolved_input_files", [])
            if item.get("status", "open") == "open"
        ]
        if unresolved_inputs:
            names = ", ".join(str(item.get("filename", "?")) for item in unresolved_inputs)
            errors.append(
                "Unsupported input files still need a recorded user decision: " + names + ". "
                "Resolve them through apply_extraction_corrections.py, then re-run Stage 1 and Stage 2."
            )
        _, support_orphans = collect_support_documents(extraction, included_units(allocation))
        if support_orphans:
            names = ", ".join(
                f"{doc.get('document_id')} ({Path(str(doc.get('source_file', ''))).name})"
                for doc in support_orphans
            )
            errors.append(
                "These supporting documents are not tied to any invoice and cannot be packaged: "
                + names + ". For each, record the invoice it backs (supports_document_id) and an "
                "optional support_type via apply_extraction_corrections.py, or exclude it with the "
                "user's reason, then re-run Stage 1 and Stage 2."
            )
    audit_roles = (("mirror_warden", "Otako - Mirror Warden"), ("gate_challenger", "Kaede - Gate Challenger"))
    audit_states = {}
    for _role, _display in audit_roles:
        _st = subagent_protocol.audit_state(_role, process_dir, allocation, allocation_path, extraction_path)
        audit_states[_role] = _st
        if (
            _st.get("current")
            and _st.get("outcome") == "block"
            and int(_st.get("blocking_count", 0) or 0) > 0
        ):
            for finding in _st.get("findings", []):
                if finding.get("severity") == "blocking":
                    errors.append(
                        f"{_display} audit blocker "
                        f"[{finding.get('code', finding.get('finding_id', '?'))}]: "
                        f"{finding.get('message', '')} Recommended action: "
                        f"{finding.get('recommended_action', '')}"
                    )
        if _st.get("current"):
            print(
                f"{_display} AUDIT: "
                f"{_st.get('outcome')} / "
                f"{_st.get('blocking_count', 0)} blocking / "
                f"{_st.get('advisory_count', 0)} advisory"
            )
        else:
            print(
                f"{_display} AUDIT: "
                f"{_st.get('status', 'missing')} - checkpoint unavailable, opted out, or stale; "
                "continuing with deterministic Stage 3 preflight only."
            )
    print_stage3_preflight_summary(allocation, errors)
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return ExitCode.COMMAND_ERROR

    units = included_units(allocation)
    proof_groups = assign_proof_numbers(units)
    mounted_support, _ = collect_support_documents(extraction, units)
    attach_support_documents_to_groups(proof_groups, mounted_support)
    rows = make_rows(units, args.requester)
    rows.sort(key=lambda r: (r["client"], r["client_charge_code"], ROW_ORDER.get(r["row_order_type"], 99), r["expense_date"], r["proof_no"]))
    annotate_event_meal_standards(rows, allocation.get("project_contexts", []))
    annotate_meal_policies(rows)
    row_text_issues = text_safety.find_suspect_text(
        text_safety.pick_fields(rows, WORKBOOK_ROW_TEXT_FIELDS),
        path="final_rows.rows",
    )
    if row_text_issues:
        print("ERROR: Refusing to write a workbook with suspect encoding damage: " + "; ".join(row_text_issues), file=sys.stderr)
        return ExitCode.COMMAND_ERROR
    meal_checks = meal_daily_cap_checks(rows)
    hotel_checks = hotel_cap_checks(rows)
    project_blocks, summary_rows = write_workbook(Path(args.output), rows, template_path, layout)
    saved_workbook_text_issues = workbook_text_issues(Path(args.output))
    if saved_workbook_text_issues:
        print("ERROR: Workbook text scan found suspect encoding damage. The workbook is not a deliverable; "
              "fix the UTF-8 answers input and re-run Stage 3. Findings: "
              + "; ".join(saved_workbook_text_issues), file=sys.stderr)
        return ExitCode.COMMAND_ERROR
    meal_check_status = aggregate_check_status(meal_checks)
    hotel_check_status = aggregate_check_status(hotel_checks)

    payload = {
        "schema_version": "final_expense_rows.v1",
        "generated_at": time_utils.iso_now(),
        "requester": args.requester,
        "source_allocation_file": str(allocation_path),
        "workbook_source": "template" if template_path else "generated",
        "template_workbook": str(template_path) if template_path else "",
        "layout_file": str(layout_path) if layout else "",
        "workbook": str(Path(args.output)),
        "proof_groups": [{k: v for k, v in group.items() if k != "units"} for group in proof_groups],
        "rows": rows,
        "rail_journey_chains": rail_chain_summaries(units),
        "expense_hint_reconciliation": allocation.get("expense_hint_reconciliation", []),
        "unresolved_expense_hint_count": sum(
            1
            for record in allocation.get("expense_hint_reconciliation", [])
            if record.get("resolution_status") not in {"not_required", "resolved"}
        ),
        "project_blocks": project_blocks,
        "summary_rows": summary_rows,
        "meal_policy_rule": meal_policy_rule(),
        "meal_daily_cap_checks": meal_checks,
        "hotel_cap_checks": hotel_checks,
        "subagent_audit": {
            "result_fingerprint": "|".join(
                f"{_role}:{audit_states[_role].get('result_fingerprint', '')}"
                for _role, _ in audit_roles
            ),
            "roles": {
                _role: subagent_protocol.review_record(audit_states[_role])
                for _role, _ in audit_roles
            },
        },
        "checks": [
            {
                "name": "meal_daily_caps",
                "caps": {
                    "business_trip_meal": money(BUSINESS_TRIP_MEAL_DAILY_CAP),
                    "local_overtime_meal": money(LOCAL_OVERTIME_MEAL_DAILY_CAP),
                },
                "classification_rule": meal_policy_rule(),
                "status": meal_check_status,
                "days_checked": len(meal_checks),
                "days_with_advisory": sum(1 for check in meal_checks if check.get("severity") == "advisory"),
                "days_requiring_confirmation": sum(1 for check in meal_checks if check["requires_user_confirmation"]),
            },
            {
                "name": "hotel_caps",
                "caps": {
                    "first_tier_city_per_night": money(FIRST_TIER_HOTEL_CAP),
                    "other_city_per_night": money(OTHER_CITY_HOTEL_CAP),
                },
                "status": hotel_check_status,
                "items_checked": len(hotel_checks),
                "items_with_advisory": sum(1 for check in hotel_checks if check.get("severity") == "advisory"),
                "items_requiring_confirmation": sum(1 for check in hotel_checks if check["requires_user_confirmation"]),
            }
        ],
    }
    # Provenance chain: packaging verifies these against the CURRENT allocation,
    # so a workbook built before later allocation edits cannot be packaged.
    payload["source_allocation_fingerprint"] = allocation.get("integrity", {}).get("fingerprint", "")
    payload["source_extraction_fingerprint"] = allocation.get("source_extraction_fingerprint", "")
    payload["blocking_policy_checks"] = len(checks_requiring_confirmation(payload))
    # Preview provenance: packaging refuses workbooks generated past open gates.
    payload["generated_with_allow_unconfirmed"] = bool(args.allow_unconfirmed)
    payload["open_allocation_questions"] = len([
        q for q in allocation.get("questions", []) if q.get("status", "open") == "open"
    ])
    workbook_path = Path(args.output)
    payload["workbook_sha256"] = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    integrity.stamp(payload, "write_reimbursement_template.py")
    write_json(process_dir / "final-expense-rows.json", payload)
    (process_dir / "final-expense-rows.md").write_text(build_markdown(payload, Path(args.output)), encoding="utf-8")
    print(f"Wrote {args.output}")
    print(f"Wrote {process_dir / 'final-expense-rows.json'}")
    print(f"Wrote {process_dir / 'final-expense-rows.md'}")
    print_meal_cap_check(meal_checks)
    print_hotel_cap_check(hotel_checks)
    print_stage3_review_summary(payload)
    if checks_requiring_confirmation(payload):
        print("ERROR: Stage 3 policy checks require applicant confirmation before final submission.", file=sys.stderr)
        print("NEXT: relay the STAGE 3 REVIEW SUMMARY above to the user VERBATIM, resolve the blocking "
              "checks via the answers updater, then RE-RUN this script. Packaging will refuse until "
              "blocking checks are zero.")
        return ExitCode.REVIEW_REQUIRED
    print("NEXT: run scripts/package_reimbursement_files.py to build the final submission package.")
    return ExitCode.SUCCESS


if __name__ == "__main__":
    raise SystemExit(main())
